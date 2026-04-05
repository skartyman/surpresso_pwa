#!/usr/bin/env python3
import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import load_workbook
except Exception as exc:
    print('[import] openpyxl is required in runtime image:', exc)
    sys.exit(1)

import subprocess

RAW_TO_INTAKE = {
    'прийнято на ремонт': 'client_repair',
    'принято на ремонт': 'client_repair',
    'приехало после аренды': 'after_rent',
    'приехало с подмены': 'after_replacement',
    'новокупленные': 'new_purchase',
    'new purchase': 'new_purchase',
}
RAW_TO_SERVICE = {
    'прийнято на ремонт': 'accepted', 'принято на ремонт': 'accepted',
    'приехало после аренды': 'accepted', 'приехало с подмены': 'accepted',
    'в роботі': 'in_progress', 'в работе': 'in_progress',
    'тест': 'testing', 'testing': 'testing',
    'готово': 'ready', 'проведено': 'processed',
    'закрыто': 'closed', 'выдано': 'closed', 'завершено': 'closed',
}
RAW_TO_COMMERCIAL = {
    'видано клієнту': 'issued_to_client', 'выдано клиенту': 'issued_to_client',
    'готово к аренде': 'ready_for_rent', 'уехало на аренду': 'out_on_rent',
    'уехало на подмену': 'out_on_replacement', 'готово к продаже': 'ready_for_sale',
    'бронь к аренде': 'reserved_for_rent', 'бронь к продаже': 'reserved_for_sale', 'продано': 'sold',
}


def key(value):
    return str(value or '').strip().lower()


def normalize(raw, mapping):
    k = key(raw)
    if not k:
        return None
    if k in mapping:
        return mapping[k]
    for source, target in mapping.items():
        if source in k:
            return target
    return None


def rows_from_sheet(ws):
    values = list(ws.values)
    if not values:
        return []
    headers = [str(x).strip() if x is not None else '' for x in values[0]]
    rows = []
    for line in values[1:]:
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            row[h] = line[i] if i < len(line) else None
        rows.append(row)
    return rows


def pick(row, *names):
    normalized = {key(k): v for k, v in row.items()}
    for name in names:
        v = normalized.get(key(name))
        if v is not None and str(v).strip() != '':
            return v
    return None


def escape(value):
    if value is None:
        return 'NULL'
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    return "'" + str(value).replace("'", "''") + "'"


def now_iso():
    return datetime.utcnow().isoformat() + 'Z'


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.getcwd(), 'Surpresso Equipment DB.xlsx')
    wb = load_workbook(xlsx_path, data_only=True)

    sheets = {name.lower(): wb[name] for name in wb.sheetnames}
    equipment_rows = rows_from_sheet(sheets.get('equipment') or wb[wb.sheetnames[0]])
    status_rows = rows_from_sheet(sheets.get('status_log') or (wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else wb[wb.sheetnames[0]]))
    photos_rows = rows_from_sheet(sheets.get('photos') or (wb[wb.sheetnames[2]] if len(wb.sheetnames) > 2 else wb[wb.sheetnames[0]]))

    imported_equipment = imported_history = imported_media = warnings = 0
    sql_statements = []

    for row in equipment_rows:
      equipment_id = str(pick(row, 'id', 'equipmentId', 'ID') or f"eq-{datetime.utcnow().timestamp()}")
      raw = str(pick(row, 'status', 'currentStatus') or '').strip() or None
      service_status = normalize(raw, RAW_TO_SERVICE)
      commercial_status = normalize(raw, RAW_TO_COMMERCIAL)
      intake_type = normalize(raw, RAW_TO_INTAKE) or 'manual_intake'
      if raw and not service_status and not commercial_status:
          warnings += 1

      sql_statements.append(f'''INSERT INTO "Equipment" ("id","clientId","type","brand","name","model","serial","internalNumber","status","ownerType","clientServiceType","equipmentType","currentStatusRaw","serviceStatus","commercialStatus","clientName","clientPhone","clientLocation","companyLocation","lastComment","folderId","folderUrl","passportPdfId","passportPdfUrl","qrUrl","isActive","createdAt","updatedAt")
VALUES ({escape(equipment_id)},{escape(pick(row,'clientId'))},{escape(str(pick(row,'legacyType','type') or 'unknown'))},{escape(str(pick(row,'brand') or 'Unknown'))},{escape(pick(row,'name'))},{escape(pick(row,'model'))},{escape(pick(row,'serial','serialNumber'))},{escape(pick(row,'internalNumber','invNumber'))},{escape(raw or 'unknown')},{escape(normalize(pick(row,'owner','ownerType'), {'company':'company','клиент':'client','client':'client','компания':'company'}))},{escape('service_contract' if str(pick(row,'isContract','contract') or '').lower() in ['1','true','yes'] else ('regular_client' if pick(row,'isContract','contract') is not None else None))},{escape(normalize(pick(row,'type','equipmentType'), {'grinder':'grinder','кофемолка':'grinder','pro':'pro_coffee','auto':'auto_coffee','filter':'filter_system'}))},{escape(raw)},{escape(service_status)},{escape(commercial_status)},{escape(pick(row,'clientName'))},{escape(pick(row,'clientPhone'))},{escape(pick(row,'clientLocation'))},{escape(pick(row,'companyLocation'))},{escape(pick(row,'comment'))},{escape(pick(row,'folderId'))},{escape(pick(row,'folderUrl'))},{escape(pick(row,'passportPdfId'))},{escape(pick(row,'passportPdfUrl'))},{escape(pick(row,'qrUrl'))},TRUE,NOW(),NOW())
ON CONFLICT ("id") DO UPDATE SET "currentStatusRaw"=EXCLUDED."currentStatusRaw","serviceStatus"=EXCLUDED."serviceStatus","commercialStatus"=EXCLUDED."commercialStatus";''')
      sql_statements.append(f'''INSERT INTO "ServiceCase" ("id","equipmentId","intakeType","serviceStatus","acceptedAt","createdAt","updatedAt")
VALUES ({escape('sc-import-' + equipment_id)},{escape(equipment_id)},{escape(intake_type)},{escape(service_status or 'accepted')},NOW(),NOW(),NOW()) ON CONFLICT DO NOTHING;''')
      imported_equipment += 1

    for row in status_rows:
      equipment_id = str(pick(row, 'equipmentId', 'id') or '').strip()
      to_raw = str(pick(row, 'toStatus', 'status', 'newStatus') or '').strip()
      if not equipment_id or not to_raw:
          continue
      from_raw = str(pick(row, 'fromStatus', 'oldStatus') or '').strip() or None
      if not normalize(to_raw, RAW_TO_SERVICE) and not normalize(to_raw, RAW_TO_COMMERCIAL):
          warnings += 1
      sql_statements.append(f'''INSERT INTO "ServiceStatusHistory" ("id","equipmentId","serviceCaseId","fromStatusRaw","toStatusRaw","fromServiceStatus","toServiceStatus","comment","actorLabel","changedAt")
VALUES ({escape('ssh-import-' + str(datetime.utcnow().timestamp()))},{escape(equipment_id)},{escape(pick(row,'serviceCaseId'))},{escape(from_raw)},{escape(to_raw)},{escape(normalize(from_raw, RAW_TO_SERVICE))},{escape(normalize(to_raw, RAW_TO_SERVICE))},{escape(pick(row,'comment'))},{escape(pick(row,'actor','changedBy'))},NOW());''')
      imported_history += 1

    for row in photos_rows:
      equipment_id = str(pick(row, 'equipmentId', 'id') or '').strip()
      url = str(pick(row, 'fileUrl', 'url', 'photoUrl') or '').strip()
      if not equipment_id or not url:
          continue
      kind = 'video' if any(url.lower().endswith(ext) for ext in ['.mp4', '.mov', '.avi']) else 'photo'
      sql_statements.append(f'''INSERT INTO "ServiceCaseMedia" ("id","equipmentId","serviceCaseId","kind","filePath","fileUrl","mimeType","originalName","fileSize","caption","createdAt")
VALUES ({escape('scm-import-' + str(datetime.utcnow().timestamp()))},{escape(equipment_id)},{escape(pick(row,'serviceCaseId'))},{escape(kind)},{escape(pick(row,'filePath') or url)},{escape(url)},{escape(pick(row,'mimeType'))},{escape(pick(row,'originalName'))},{escape(int(pick(row,'fileSize','size') or 0))},{escape(pick(row,'caption'))},NOW());''')
      imported_media += 1

    sql = 'BEGIN;\n' + '\n'.join(sql_statements) + '\nCOMMIT;'
    temp_sql = '/tmp/surpresso_import_service_equipment.sql'
    with open(temp_sql, 'w', encoding='utf-8') as f:
        f.write(sql)

    database_url = os.environ.get('DATABASE_URL')
    if database_url:
        subprocess.run(['psql', database_url, '-f', temp_sql], check=True)

    print('[import] import completed summary', json.dumps({
        'importedEquipment': imported_equipment,
        'importedHistory': imported_history,
        'importedMedia': imported_media,
        'migrationWarnings': warnings,
        'completedAt': now_iso(),
    }, ensure_ascii=False))


if __name__ == '__main__':
    main()
